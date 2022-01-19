import openpyxl
import time
import subprocess
import itertools

wb = openpyxl.load_workbook('sheet.xlsx')
#print(wb.sheetnames)
sheet = wb['Data']


def remove(sheet):                                              #function to remove empty rows in the middle

    for row in sheet.iter_rows():

        # all() return False if all of the row value is None
        if not all(cell.value for cell in row):
            # detele the empty row
            sheet.delete_rows(row[0].row, 1)

            # recursively call the remove() with modified sheet data
            remove(sheet)

            return

#print('max row is', sheet.max_row)

#print evetything
#for i in range(1, sheet.max_row+1, 1):
#        print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value, sheet.cell(row=i, column=5).value)

#count of email IDs in dictionary

email_ids = list()
counts = dict()

for i in range(2, sheet.max_row+1, 1):
    a = sheet.cell(row=i, column=5).value
    email_ids.append(a)

#print(email_ids)

for email_id in email_ids:
    counts[email_id] = counts.get(email_id, 0) + 1
#print(counts)

unique_mail_ids = list(counts)
#print(unique_mail_ids)
total_number_of_unique_mail_ids = len(unique_mail_ids)
#print(total_number_of_unique_mail_ids)
#print(unique_mail_ids)
total = total_number_of_unique_mail_ids

for i in range(0, total ):                          #to slice workbook into multiple sheets per owner name
    wb.create_sheet(index=i+1, title=str(i))
    #time.sleep(1)
    #print(i, unique_mail_ids[i])

wb.save('sheet.xlsx')

z = 0

#current_name = unique_mail_ids[0]

for j in unique_mail_ids:                              #write data to all the sheets in the workbook


    for i in range(2, sheet.max_row+1, 1):
        #a = sheet.cell(row=i, column=5).value
        if j == sheet.cell(row=i, column=5).value:
            #print(sheet.cell(row=i, column= 1).value, sheet.cell(row=i, column= 2).value, sheet.cell(row=i, column= 3).value, sheet.cell(row=i, column= 4).value, sheet.cell(row=i, column= 5).value)
            a = sheet.cell(row=i, column=1).value
            b = sheet.cell(row=i, column=2).value
            c = sheet.cell(row=i, column=3).value
            d = sheet.cell(row=i, column=4).value
            e = sheet.cell(row=i, column=5).value
           # f = sheet.cell(row=i, column=6).value
            x = str(z)
            sheet1 = wb[x]

            sheet1['A1'] = 'Region'
            sheet1['B1'] = 'Server Name'
            sheet1['C1'] = 'GXP'
            sheet1['D1'] = 'Usage as per CMDB'
            sheet1['E1'] = 'Owner1'
           # sheet1['F1'] = 'Owner2'



            sheet1.cell(row=i, column=1).value = a
            sheet1.cell(row=i, column=2).value = b
            sheet1.cell(row=i, column=3).value = c
            sheet1.cell(row=i, column=4).value = d
            sheet1.cell(row=i, column=5).value = e
           # sheet1.cell(row=i, column=6).value = f

        sheet = wb['Data']
    z = z + 1
            #print('got the desired value at row', i, sheet.cell(row=i, column=5).value)


wb.save('sheet.xlsx')


for j in range(0, total, 1):           # calling remove() to remove empty rows in the middle
    h = str(j)
    sheet1 = wb[h]
    for row in sheet1:
        remove(sheet1)

wb.save('sheet.xlsx')


wb = openpyxl.load_workbook('sheet.xlsx')

for i in range(0, total, 1):
    k = str(i)
    sheet2 = wb[k]
    owner1_name = sheet2.cell(2, 5).value

    row = sheet2.max_row
    # second_owner = sheet2.cell(2, 6).value
    column = sheet2.max_column
    server_list = list()
    usage_list = list()
    for j in range(2, row + 1):
        servers = sheet2.cell(j, 2).value
        usage = sheet2.cell(j, 4).value
        # actual_list_with_region = str(servers) + "             " + str(region)
        server_list.append(servers)
        usage_list.append(usage)

    #print(server_list)
    #print(usage_list)

    # print(server_list)

    file1 = open("mailcontent.txt", "w")

    L2 = ["""<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office"
    xmlns:w="urn:schemas-microsoft-com:office:word" xmlns:m="http://schemas.microsoft.com/office/2004/12/omml"
    xmlns="http://www.w3.org/TR/REC-html40">

<head>
    <META HTTP-EQUIV="Content-Type" CONTENT="text/html; charset=us-ascii">
    <meta name=Generator content="Microsoft Word 15 (filtered medium)">
    <style>
        <!--
        /* Font Definitions */
        @font-face {
            font-family: "Cambria Math";
            panose-1: 2 4 5 3 5 4 6 3 2 4;
        }

        @font-face {
            font-family: Calibri;
            panose-1: 2 15 5 2 2 2 4 3 2 4;
        }

        @font-face {
            font-family: "Segoe UI";
            panose-1: 2 11 5 2 4 2 4 2 2 3;
        }

        /* Style Definitions */
        p.MsoNormal,
        li.MsoNormal,
        div.MsoNormal {
            margin: 0cm;
            margin-bottom: .0001pt;
            font-size: 11.0pt;
            font-family: "Calibri", sans-serif;
        }

        a:link,
        span.MsoHyperlink {
            mso-style-priority: 99;
            color: #0563C1;
            text-decoration: underline;
        }

        a:visited,
        span.MsoHyperlinkFollowed {
            mso-style-priority: 99;
            color: #954F72;
            text-decoration: underline;
        }

        span.EmailStyle17 {
            mso-style-type: personal-compose;
            font-family: "Calibri", sans-serif;
            color: windowtext;
        }

        .MsoChpDefault {
            mso-style-type: export-only;
            mso-fareast-language: EN-US;
        }

        @page WordSection1 {
            size: 612.0pt 792.0pt;
            margin: 72.0pt 72.0pt 72.0pt 72.0pt;
        }

        div.WordSection1 {
            page: WordSection1;
        }
        -->
    </style>
    <!--[if gte mso 9]><xml>
    <o:shapedefaults v:ext="edit" spidmax="1026" />
    </xml><![endif]-->
    <!--[if gte mso 9]><xml>
    <o:shapelayout v:ext="edit">
    <o:idmap v:ext="edit" data="1" />
    </o:shapelayout></xml><![endif]-->
</head>

<body lang=EN-IN link="#0563C1" vlink="#954F72">
    <div class=WordSection1>
        <p class=MsoNormal><b><span lang=EN-US>Dear Linux Customer</span></b><span lang=EN-US>,<o:p></o:p></span></p>
        <p class=MsoNormal><span lang=EN-US>
                <o:p>&nbsp;</o:p>
            </span></p>
        <p class=MsoNormal><span lang=EN-US>
                <o:p>&nbsp;</o:p>
            </span></p>
        <p class=MsoNormal><span lang=EN-US>Kindly provide 2 hours downtime window to patch the below server(s):<o:p>
                </o:p></span></p>
        <p class=MsoNormal><span lang=EN-US>
                <o:p>&nbsp;</o:p>
            </span></p>
        <p class=MsoNormal><span lang=EN-US>
                <o:p>&nbsp;</o:p>
            </span></p>
        <table class=MsoNormalTable border=0 cellspacing=0 cellpadding=0 width=0
            style='width:184.0pt;border-collapse:collapse'>
            <tr style='height:14.5pt'>
                <td width=128 nowrap
                    style='width:96.0pt;border:solid windowtext 1.0pt;background:#002060;padding:0cm 5.4pt 0cm 5.4pt;height:14.5pt'>
                    <p class=MsoNormal><b><span style='font-size:10.0pt;color:white'>Server Name</span></b><span
                            style='color:white'> </span><b><span style='font-size:10.0pt;color:white'>
                                <o:p></o:p>
                            </span></b></p>
                </td>
                <td width=117 nowrap
                    style='width:88.0pt;border:solid windowtext 1.0pt;border-left:none;background:#002060;padding:0cm 5.4pt 0cm 5.4pt;height:14.5pt'>
                    <p class=MsoNormal><b><span style='font-size:10.0pt;color:white'>Usage as per CMDB <o:p></o:p>
                                </span></b></p>
                </td>
            </tr>"""]






    L3 = ["""</table>
        <p class=MsoNormal><span style='font-family:"Cambria",serif;color:#002060'>
                <o:p>&nbsp;</o:p>
            </span></p>
        <p class=MsoNormal><span style='font-family:"Cambria",serif;color:#002060'>
                <o:p>&nbsp;</o:p>
            </span></p>
        <p class=MsoNormal><span style='font-family:"Cambria",serif;color:#002060'>Regards,<o:p></o:p></span></p>
        <p class=MsoNormal><span style='font-family:"Cambria",serif;color:#002060'>Linux Team<o:p></o:p></span></p>
    </div>
</body>

</html>"""]

    file1.writelines(L2)
    # file1.write("Hello \n")

    file1.close()

    file1 = open("mailcontent.txt", "a+")

    for (h, n) in zip(server_list, usage_list):
        L11 = ["""<tr style='height:14.5pt'>
                <td width=156 nowrap valign=bottom
                    style='width:117.0pt;border:solid windowtext 1.0pt;border-top:none;padding:0cm 5.4pt 0cm 5.4pt;height:14.5pt'>
                    <p class=MsoNormal><span style='color:black;mso-fareast-language:EN-IN'>"""]
        file1.writelines(L11)
        file1.writelines(h)

        L12 = ["""<o:p></o:p>
                            </span></p>
                </td>
                <td width=117 nowrap valign=bottom
                    style='width:88.0pt;border-top:none;border-left:none;border-bottom:solid windowtext 1.0pt;border-right:solid windowtext 1.0pt;padding:0cm 5.4pt 0cm 5.4pt;height:14.5pt'>
                    <p class=MsoNormal><span style='color:black;mso-fareast-language:EN-IN'>"""]
        file1.writelines(L12)
        file1.writelines(n)



        #print(h, n)

        L13 = ["""<o:p></o:p></span></p>
                </td>
            </tr>"""]
        file1.writelines(L13)


#    with open('mailcontent.txt', 'a') as f:
#        f.writelines('\n'.join(server_list))

    file1.close()

    file1 = open("mailcontent.txt", "a+")

    file1.writelines(L3)

    file1.close()

    l = """mailx -s "To reschedule patching of PROD server" -a "From: Unix Patch Management <unix-patch-management@xxx.com>" -a "Content-type: text/html" -c unix-patch-management@xxx.com,naveen@gmail.com. """
    m = str(owner1_name)
    # o = ","
    # p = second_owner
    n = """ < "mailcontent.txt" """

    # cmd = l + m + o + p + n
    cmd = l + m + n
    # print(cmd)

    # cmd = """mailx -s "Linux OS Patching and Downtime Request - Announcemen" -a "From: CCP Linux Team<ccpbayerops.in@capgemini.com>" -c shreyas.p@capgemini.com. shreyas.p@capgemini.com < "mailcontent.txt" """

    subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
    print("mail sent to:")
    print(i)
    print("Delay of 15 seconds")
    time.sleep(15)

######################################################################  END OF PROGRAM  #############################################################################

